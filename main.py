from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from typing import List
import httpx
import pandas as pd
import asyncio
import os
import uuid
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = FastAPI(title="EPİAŞ Veri Çekme API", version="3.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# 🔧 GENEL AYARLAR
# ==========================================
USERNAME = "beratr10ltd@gmail.com"
PASSWORD = "Beroto11."

URL_TGT         = "https://giris.epias.com.tr/cas/v1/tickets"
URL_ORG         = "https://seffaflik.epias.com.tr/electricity-service/v1/generation/data/organization-list"
URL_UEVCB       = "https://seffaflik.epias.com.tr/electricity-service/v1/generation/data/uevcb-list-bulk"
URL_KGUP        = "https://seffaflik.epias.com.tr/electricity-service/v1/generation/data/dpp-bulk"
URL_GOP_ESLESME = "https://seffaflik.epias.com.tr/electricity-service/v1/markets/dam/data/clearing-quantity"
URL_GIP_ESLESME = "https://seffaflik.epias.com.tr/electricity-service/v1/markets/idm/data/matching-quantity"

KGUP_GRUP_HARITASI = {
    "İzmir":        ["İZMİR DGKÇS_GR1(Doğalgaz)", "İZMİR DGKÇS_GR2(Doğalgaz)"],
    "ADAPAZARI":    ["ADAPAZARI DGKÇS"],
    "GEBZE":        ["GEBZE DGKÇS_1", "GEBZE DGKÇS_-2"],
    "İÇ ANADOLU":  ["İÇ ANADOLU DGKÇS"],
    "RWE":          ["RWE_TURCAS_GUNEY"],
    "BANDIRMA 1":   ["ENERJİSA BANDIRMA SANTRALI"],
    "BANDIRMA 2":   ["BANDIRMA II DGKÇS"],
    "HAMİTABAT":    ["HAMİTABAT ÜNİTE-10", "HAMİTABAT ÜNİTE-20"],
    "İSKEN":        ["İSKENDERUN İTHAL KÖMÜR SANTRALI-2", "İSKENDERUN İTHAL KÖMÜR SANTRALI-1"],
    "CENAL":        ["CENAL TES(TR1+TRA)", "CENAL TES(TR2)"],
    "EMBA":         ["HUNUTLU TES_TR1", "HUNUTLU TES_TR2"],
    "ATLAS":        ["ATLAS TES-A","ATLAS TES-B"],
    "EREN":         ["ZETES 1", "ZETES 2-A", "ZETES 2-B", "ZETES 3-A", "ZETES 3-B"],
    "AKSA ANTALYA": ["ANTALYA ENERJİ SANTRALİ(Doğalgaz)"],
    "EGEMER":       ["Erzin Santrali"],
    "ACWA":         ["ACWA POWER KIRIKKALE DGKÇS"],
    "CENGİZ":       ["CENGİZ 610 DGKÇS"],
    "BİLGİN":       ["BİLGİN SAMSUN DGKÇS PT1(Doğalgaz)", "BİLGİN SAMSUN DGKÇS PT2(Doğalgaz)"],
    "BAYMİNA":      ["ANKARA DGKÇS"],
    "HABAŞ":        ["ALİAĞA DGKÇS(Habaş DGKÇS/DB-1)"]
}

GOP_ORGANIZASYONLAR = [
    {"organizationId": 19860,  "shortName": "HUB"},
    {"organizationId": 20700,  "shortName": "MERK"},
    {"organizationId": 100200, "shortName": "ALTAİR"},
    {"organizationId": 20300,  "shortName": "RKARE"},
    {"organizationId": 15220,  "shortName": "NEXUS"},
]

OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)
jobs: dict = {}

# ==========================================
# 🛠️ YARDIMCI FONKSİYONLAR
# ==========================================

async def get_tgt_async(client: httpx.AsyncClient) -> str:
    res = await client.post(
        URL_TGT,
        headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "text/plain"},
        data={"username": USERNAME, "password": PASSWORD},
        timeout=30
    )
    if res.status_code not in (200, 201):
        raise Exception(f"TGT alınamadı: {res.status_code}")
    return res.text

def api_headers(tgt: str) -> dict:
    return {"TGT": tgt, "Content-Type": "application/json", "Accept": "application/json"}

def get_dates():
    bugun = datetime.now()
    yarin = bugun + timedelta(days=1)
    return {
        "tomorrow_str":  yarin.strftime("%Y-%m-%d"),
        "tomorrow_full": yarin.strftime("%Y-%m-%dT00:00:00+03:00"),
        "start_iso":     bugun.strftime("%Y-%m-%dT00:00:00+03:00"),
        "end_iso":       yarin.strftime("%Y-%m-%dT00:00:00+03:00"),
    }

def update_job(job_id, status, message, progress=None, file=None):
    jobs[job_id] = {"status": status, "message": message, "progress": progress, "file": file}

def style_xlsx_basic(file_path: str, col_width: int = 15):
    try:
        wb = load_workbook(file_path)
        for s in wb.sheetnames:
            ws = wb[s]
            for c in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(c)].width = col_width
        wb.save(file_path)
    except:
        pass

# ==========================================
# 🏭 KGÜP — Tam Async + Paralel Batch
# ==========================================

async def process_kgup_async(job_id: str):
    update_job(job_id, "running", "TGT alınıyor...", 5)

    async with httpx.AsyncClient(timeout=60) as client:
        try:
            tgt = await get_tgt_async(client)
        except Exception as e:
            update_job(job_id, "error", str(e)); return

        headers = api_headers(tgt)
        dates   = get_dates()

        # Adım 1: Organizasyonlar
        update_job(job_id, "running", "Organizasyonlar çekiliyor...", 10)
        try:
            res = await client.post(URL_ORG, headers=headers,
                                    json={"startDate": dates["start_iso"], "endDate": dates["end_iso"]})
            if res.status_code != 200:
                update_job(job_id, "error", "Organizasyon listesi alınamadı."); return
            org_ids = [item["organizationId"] for item in res.json().get("items", []) if "organizationId" in item]
        except Exception as e:
            update_job(job_id, "error", f"Org hatası: {e}"); return

        update_job(job_id, "running", f"{len(org_ids)} org bulundu, UEVÇB çekiliyor...", 20)

        # Adım 2: UEVÇB — Paralel
        async def fetch_uevcb(batch):
            try:
                r = await client.post(URL_UEVCB, headers=headers,
                                      json={"organizationIds": batch, "startDate": dates["tomorrow_full"]})
                if r.status_code == 200:
                    return [item["id"] for item in r.json().get("items", []) if "id" in item]
            except:
                pass
            return []

        uevcb_batches = [org_ids[i:i+500] for i in range(0, len(org_ids), 500)]
        uevcb_results = await asyncio.gather(*[fetch_uevcb(b) for b in uevcb_batches])
        uevcb_ids     = [uid for r in uevcb_results for uid in r]

        update_job(job_id, "running", f"{len(uevcb_ids)} UEVÇB bulundu, KGÜP çekiliyor...", 35)

        # Adım 3: KGÜP — Paralel + Semaphore
        kgup_batches  = [uevcb_ids[i:i+900] for i in range(0, len(uevcb_ids), 900)]
        total_b       = len(kgup_batches)
        final_veriler = []
        sem           = asyncio.Semaphore(5)

        async def fetch_kgup(batch, idx):
            async with sem:
                for _ in range(3):
                    try:
                        r = await client.post(URL_KGUP, headers=headers,
                                              json={"uevcbIds": batch, "date": dates["tomorrow_full"], "region": "TR1"})
                        if r.status_code == 200:
                            update_job(job_id, "running", f"Batch {idx+1}/{total_b} işlendi",
                                       35 + int(((idx+1)/total_b)*45))
                            return [{"time": i.get("time"), "uevcbName": i.get("uevcbName"), "toplam": i.get("toplam")}
                                    for i in r.json().get("items", [])]
                        elif r.status_code == 401:
                            update_job(job_id, "error", "TGT süresi doldu."); return []
                    except:
                        pass
                    await asyncio.sleep(3)
                return []

        kgup_results = await asyncio.gather(*[fetch_kgup(b, i) for i, b in enumerate(kgup_batches)])
        for r in kgup_results:
            final_veriler.extend(r)

    if not final_veriler:
        update_job(job_id, "error", "KGÜP verisi bulunamadı."); return

    update_job(job_id, "running", "Excel oluşturuluyor...", 85)
    df = pd.DataFrame(final_veriler)
    df["toplam"] = pd.to_numeric(df["toplam"]).fillna(0)

    pivot_detay = df.pivot_table(index="time", columns="uevcbName", values="toplam",
                                  aggfunc="sum", fill_value=0).sort_index()
    pivot_ozet = pd.DataFrame(index=pivot_detay.index)
    for grup, liste in KGUP_GRUP_HARITASI.items():
        cols = [c for c in liste if c in pivot_detay.columns]
        if cols:
            pivot_ozet[grup] = pivot_detay[cols].sum(axis=1)

    file_name = f"KGUP_{dates['tomorrow_str']}_{job_id[:8]}.xlsx"
    file_path = os.path.join(OUTPUT_DIR, file_name)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        pivot_detay.to_excel(writer, sheet_name=f"{dates['tomorrow_str']}_kgüp")
        pivot_ozet.to_excel(writer, sheet_name="Özet")

    style_xlsx_basic(file_path)
    update_job(job_id, "done", "KGÜP raporu hazır!", 100, file_name)


def run_kgup(job_id: str):
    asyncio.run(process_kgup_async(job_id))


# ==========================================
# 📊 GÖP EŞLEŞME — Org Bazlı Döngü
# ==========================================

class GopEslesmeRequest(BaseModel):
    start_date: str = Field(..., example="2024-01-01")
    end_date:   str = Field(..., example="2024-01-31")


async def fetch_with_retry(client, url, headers, payload, retries=3, delay=2):
    """Verilen URL'ye retry mekanizmasıyla POST atar, items listesini döner."""
    for _ in range(retries):
        try:
            res = await client.post(url, headers=headers, json=payload)
            if res.status_code == 200:
                return res.json().get("items", [])
            elif res.status_code == 401:
                raise Exception("TGT_EXPIRED")
            await asyncio.sleep(delay)
        except Exception as e:
            if "TGT_EXPIRED" in str(e):
                raise
            await asyncio.sleep(delay)
    return []


def parse_gop(items: list) -> pd.DataFrame:
    """GÖP items → (Tarih, Saat) indexli GÖP Eşleşme sütunu."""
    if not items:
        return pd.DataFrame(columns=["Tarih", "Saat", "GÖP Eşleşme (MWh)"])
    df = pd.DataFrame(items)
    df["Tarih"] = pd.to_datetime(df["date"]).dt.strftime("%Y-%m-%d")
    df["Saat"]  = df["hour"]
    df["GÖP Eşleşme (MWh)"] = pd.to_numeric(df["matchedBids"], errors="coerce").fillna(0) \
                             - pd.to_numeric(df["matchedOffers"], errors="coerce").fillna(0)
    return df[["Tarih", "Saat", "GÖP Eşleşme (MWh)"]].copy()


def parse_gip(items: list) -> pd.DataFrame:
    """GİP items → (Tarih, Saat) indexli GİP Eşleşme sütunu.
    kontratAdi örnek: 'PH26040300'  → tarih=2026-04-03, saat=00:00
    """
    if not items:
        return pd.DataFrame(columns=["Tarih", "Saat", "GİP Eşleşme (MWh)"])

    rows = []
    for item in items:
        if item.get("kontratTuru") != "Saatlik":
            continue
        kod = item.get("kontratAdi", "")
        if len(kod) < 10:
            continue
        try:
            # YYMMDD
            yy, mm, dd = kod[2:4], kod[4:6], kod[6:8]
            tarih = f"20{yy}-{mm}-{dd}"
            saat  = f"{kod[8:10]}:00"
        except:
            continue
        bid = pd.to_numeric(item.get("clearingQuantityBid", 0), errors="coerce") or 0
        ask = pd.to_numeric(item.get("clearingQuantityAsk", 0), errors="coerce") or 0
        rows.append({"Tarih": tarih, "Saat": saat, "GİP Eşleşme (MWh)": bid - ask})

    if not rows:
        return pd.DataFrame(columns=["Tarih", "Saat", "GİP Eşleşme (MWh)"])

    df = pd.DataFrame(rows)
    # Aynı tarih+saat'te birden fazla kontrat olabilir → topla
    df = df.groupby(["Tarih", "Saat"], as_index=False)["GİP Eşleşme (MWh)"].sum()
    return df


async def process_gop_eslesme_async(job_id: str, start_date: str, end_date: str):
    update_job(job_id, "running", "TGT alınıyor...", 5)
    total = len(GOP_ORGANIZASYONLAR)

    async with httpx.AsyncClient(timeout=30) as client:
        try:
            tgt = await get_tgt_async(client)
        except Exception as e:
            update_job(job_id, "error", f"TGT hatası: {e}"); return

        hdrs         = api_headers(tgt)
        org_verileri = {}   # shortName → {"gop": df, "gip": df, "has_data": bool}

        for idx, org in enumerate(GOP_ORGANIZASYONLAR):
            org_id = org["organizationId"]
            short  = org["shortName"]
            prog   = 10 + int((idx / total) * 74)
            update_job(job_id, "running", f"{short} — GÖP + GİP çekiliyor... ({idx+1}/{total})", prog)

            payload = {
                "organizationId": org_id,
                "startDate":      f"{start_date}T00:00:00+03:00",
                "endDate":        f"{end_date}T00:00:00+03:00",
            }

            try:
                gop_items, gip_items = await asyncio.gather(
                    fetch_with_retry(client, URL_GOP_ESLESME, hdrs, payload),
                    fetch_with_retry(client, URL_GIP_ESLESME, hdrs, payload),
                )
            except Exception as e:
                if "TGT_EXPIRED" in str(e):
                    update_job(job_id, "error", "TGT süresi doldu."); return
                gop_items, gip_items = [], []

            df_gop   = parse_gop(gop_items)
            df_gip   = parse_gip(gip_items)
            has_data = not (df_gop.empty and df_gip.empty)
            org_verileri[short] = {"gop": df_gop, "gip": df_gip, "has_data": has_data}

    update_job(job_id, "running", "Excel oluşturuluyor...", 88)

    # ── Tüm Tarih+Saat kombinasyonlarını topla ──
    all_keys = set()
    for v in org_verileri.values():
        for df in [v["gop"], v["gip"]]:
            if not df.empty and "Tarih" in df.columns and "Saat" in df.columns:
                for _, row in df.iterrows():
                    all_keys.add((row["Tarih"], row["Saat"]))
    all_keys = sorted(all_keys)

    file_name = f"Genel_Eslesme_{start_date}_{end_date}_{job_id[:8]}.xlsx"
    file_path = os.path.join(OUTPUT_DIR, file_name)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Genel Eşleşme"

    orgs   = list(org_verileri.keys())
    n_orgs = len(orgs)

    # ── Stil sabitleri ──
    thin      = Side(style="thin", color="2D2D44")
    border    = Border(bottom=thin, right=thin)
    fill_dark  = PatternFill("solid", fgColor="0D0D20")
    fill_head  = PatternFill("solid", fgColor="1E1E2E")
    fill_title = PatternFill("solid", fgColor="0A0A1A")
    fill_green = PatternFill("solid", fgColor="002B14")
    fill_red   = PatternFill("solid", fgColor="2B0008")
    fill_empty = PatternFill("solid", fgColor="151520")
    font_green  = Font(name="Calibri", size=10, bold=True, color="00FF88")
    font_red    = Font(name="Calibri", size=10, bold=True, color="FF3366")
    font_normal = Font(name="Calibri", size=10, color="E2E8F0")
    font_empty  = Font(name="Calibri", size=10, italic=True, color="4A5568")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # ── SATIR 1: Org başlıkları (A1:B1 boş, sonra her org 2 sütun) ──
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:B1")
    ws["A1"].fill = fill_title
    ws["A1"].border = border

    for i, short in enumerate(orgs):
        cs  = get_column_letter(3 + i * 2)
        ce  = get_column_letter(4 + i * 2)
        ws.merge_cells(f"{cs}1:{ce}1")
        c           = ws[f"{cs}1"]
        c.value     = short
        c.font      = Font(name="Calibri", bold=True, size=16, color="00E5FF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = fill_title
        c.border    = border

    # ── SATIR 2: Sütun başlıkları ──
    ws.row_dimensions[2].height = 20
    for col, txt in [(1, "Tarih"), (2, "Saat")]:
        c           = ws.cell(row=2, column=col)
        c.value     = txt
        c.font      = font_header
        c.fill      = fill_head
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border

    for i in range(n_orgs):
        for offset, label in [(0, "GÖP Eşleşme (MWh)"), (1, "GİP Eşleşme (MWh)")]:
            c           = ws.cell(row=2, column=3 + i * 2 + offset)
            c.value     = label
            c.font      = font_header
            c.fill      = fill_head
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border

    # ── Lookup tabloları ──
    def make_lookup(df, val_col):
        if df.empty or val_col not in df.columns:
            return {}
        return {(r["Tarih"], r["Saat"]): r[val_col] for _, r in df.iterrows()}

    org_lookups = {
        short: {
            "gop":      make_lookup(v["gop"], "GÖP Eşleşme (MWh)"),
            "gip":      make_lookup(v["gip"], "GİP Eşleşme (MWh)"),
            "has_data": v["has_data"],
        }
        for short, v in org_verileri.items()
    }

    # ── SATIRLAR 3+: Veri ──
    for row_idx, (tarih, saat) in enumerate(all_keys, start=3):
        ws.row_dimensions[row_idx].height = 16

        for col, val in [(1, tarih), (2, saat)]:
            c           = ws.cell(row=row_idx, column=col, value=val)
            c.font      = font_normal
            c.fill      = fill_dark
            c.alignment = Alignment(horizontal="center")
            c.border    = border

        for i, short in enumerate(orgs):
            lk = org_lookups[short]
            for offset, key in [(0, "gop"), (1, "gip")]:
                c           = ws.cell(row=row_idx, column=3 + i * 2 + offset)
                c.alignment = Alignment(horizontal="center")
                c.border    = border

                if not lk["has_data"]:
                    c.value = "Veri Yok"; c.font = font_empty; c.fill = fill_empty
                else:
                    val = lk[key].get((tarih, saat), None)
                    if val is None:
                        c.value = "Veri Yok"; c.font = font_empty; c.fill = fill_empty
                    else:
                        c.value          = val
                        c.number_format  = '+0.00;-0.00;0.00'
                        if val > 0:
                            c.font = font_green; c.fill = fill_green
                        elif val < 0:
                            c.font = font_red;   c.fill = fill_red
                        else:
                            c.font = font_normal; c.fill = fill_dark

    # ── Sütun genişlikleri ──
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 8
    for i in range(n_orgs):
        ws.column_dimensions[get_column_letter(3 + i * 2)].width     = 22
        ws.column_dimensions[get_column_letter(4 + i * 2)].width = 22

    ws.freeze_panes = "C3"
    wb.save(file_path)
    update_job(job_id, "done", "Genel Eşleşme raporu hazır!", 100, file_name)


def run_gop_eslesme(job_id: str, start_date: str, end_date: str):
    asyncio.run(process_gop_eslesme_async(job_id, start_date, end_date))


# ==========================================
# 🌐 API ENDPOINTLER
# ==========================================

@app.get("/")
def root():
    return FileResponse("static/index.html")

@app.post("/api/kgup/start")
def start_kgup(background_tasks: BackgroundTasks):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "message": "Kuyruğa alındı", "progress": 0, "file": None}
    background_tasks.add_task(run_kgup, job_id)
    return {"job_id": job_id}

@app.post("/api/gop-eslesme/start")
def start_gop_eslesme(req: GopEslesmeRequest, background_tasks: BackgroundTasks):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "message": "Kuyruğa alındı", "progress": 0, "file": None}
    background_tasks.add_task(run_gop_eslesme, job_id, req.start_date, req.end_date)
    return {"job_id": job_id}

@app.get("/api/job/{job_id}")
def get_job_status(job_id: str):
    job = jobs.get(job_id)
    if not job:
        return JSONResponse(status_code=404, content={"error": "İş bulunamadı"})
    return job

@app.get("/api/download/{filename}")
def download_file(filename: str):
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        return JSONResponse(status_code=404, content={"error": "Dosya bulunamadı"})
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )


app.mount("/static", StaticFiles(directory="static"), name="static")
